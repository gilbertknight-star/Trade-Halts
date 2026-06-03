"""
generate_report.py — Weekly Excel report generator.
=====================================================

Reads reports/validator_log.csv (written daily by daily_validator.py)
and produces reports/halt_trader_report.xlsx with three sheets:

  Daily Log   — one row per halt event across all days
  Summary     — one row per trading day (totals, match rate, slippage)
  Equity      — running account balance over time

Usage (run any time, typically end of week):
  python bot/generate_report.py

Output:
  reports/halt_trader_report.xlsx

Then on Windows:
  scp trade-server:/root/Live_Trader_Halts/bot/reports/halt_trader_report.xlsx .
"""

import csv
import sys
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import (
        Alignment, Border, Font, GradientFill,
        PatternFill, Side,
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import LineChart, Reference
    from openpyxl.chart.series import SeriesLabel
except ImportError:
    print("openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

REPORT_DIR   = Path("reports")
CSV_PATH     = REPORT_DIR / "validator_log.csv"
OUTPUT_PATH  = REPORT_DIR / "halt_trader_report.xlsx"

# ── Colour palette ────────────────────────────────────────────────────────────
C_HEADER_BG   = "1E3A5F"   # dark navy
C_HEADER_FG   = "FFFFFF"
C_PASS_BG     = "D1FAE5"   # light green
C_FAIL_BG     = "FEF3C7"   # light yellow
C_MISS_BG     = "FEE2E2"   # light red
C_MATCH_BG    = "DBEAFE"   # light blue
C_ALT_ROW     = "F8FAFC"   # very light grey
C_WHITE       = "FFFFFF"
C_BORDER      = "CBD5E1"
C_POS         = "16A34A"   # green text
C_NEG         = "DC2626"   # red text
C_NEUTRAL     = "334155"   # dark slate

THIN = Side(style="thin", color=C_BORDER)
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# ── Helpers ───────────────────────────────────────────────────────────────────

def header_cell(ws, row, col, value, width=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(bold=True, color=C_HEADER_FG, size=10)
    c.fill      = PatternFill("solid", fgColor=C_HEADER_BG)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border    = BORDER
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width
    return c


def data_cell(ws, row, col, value, bold=False, color=C_NEUTRAL,
              bg=None, fmt=None, align="left"):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(bold=bold, color=color, size=10)
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border    = BORDER
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    if fmt:
        c.number_format = fmt
    return c


def safe_float(v, default=None):
    try:
        return float(v) if v not in (None, "", "None") else default
    except (ValueError, TypeError):
        return default


def safe_int(v, default=0):
    try:
        return int(float(v)) if v not in (None, "", "None") else default
    except (ValueError, TypeError):
        return default


# ── Load CSV ──────────────────────────────────────────────────────────────────

def load_csv() -> list[dict]:
    if not CSV_PATH.exists():
        print(f"No data found at {CSV_PATH}")
        print("Run daily_validator.py at least once first.")
        sys.exit(1)
    with open(CSV_PATH, newline="") as f:
        return list(csv.DictReader(f))


# ── Sheet 1: Daily Log ────────────────────────────────────────────────────────

def build_daily_log(wb, rows: list[dict]) -> None:
    ws = wb.create_sheet("Daily Log")
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 32

    headers = [
        ("Date",          11),
        ("Symbol",         9),
        ("Resume",         9),
        ("RVOL",           7),
        ("Gap-up",         8),
        ("Filter",         8),
        ("Reason",        28),
        ("Sim Entry",      9),
        ("Sim Exit",       9),
        ("Sim P&L",        9),
        ("Sim Shares",     9),
        ("Live Action",   10),
        ("Actual Entry",  10),
        ("Actual Exit",   10),
        ("Actual P&L",    10),
        ("Slip/Share",    10),
        ("Slip Total",    10),
    ]

    for col, (label, width) in enumerate(headers, 1):
        header_cell(ws, 1, col, label, width)

    for i, row in enumerate(rows, 2):
        bg = C_ALT_ROW if i % 2 == 0 else C_WHITE

        filt        = row.get("filter_result", "")
        live_action = row.get("live_bot_action", "")
        sim_pnl     = safe_float(row.get("sim_pnl"))
        act_pnl     = safe_float(row.get("actual_pnl"))
        slip_total  = safe_float(row.get("entry_slip_total"))
        rvol        = safe_float(row.get("rvol"))
        up_move     = safe_float(row.get("up_move"))
        slip_ps     = safe_float(row.get("entry_slip_per_sh"))

        # Row highlight based on outcome
        if filt == "PASS" and live_action == "TRADED":
            row_bg = C_MATCH_BG
        elif filt == "PASS" and live_action == "SKIPPED":
            row_bg = C_MISS_BG
        elif filt == "SKIP" and live_action == "TRADED":
            row_bg = "FFF7ED"   # orange tint — unexpected
        else:
            row_bg = bg

        data_cell(ws, i,  1, row.get("date"),        bg=row_bg, align="center")
        data_cell(ws, i,  2, row.get("symbol"),       bg=row_bg, bold=True)
        data_cell(ws, i,  3, row.get("resume_time"),  bg=row_bg, align="center")
        data_cell(ws, i,  4, rvol,                    bg=row_bg, align="right",
                  fmt="0.00")
        data_cell(ws, i,  5, up_move,                 bg=row_bg, align="right",
                  fmt="0.00%")
        # Filter badge
        filt_color = C_POS if filt == "PASS" else C_NEUTRAL
        data_cell(ws, i,  6, filt, bg=row_bg, bold=(filt == "PASS"),
                  color=filt_color, align="center")
        data_cell(ws, i,  7, row.get("filter_reason"), bg=row_bg, color="64748B")
        data_cell(ws, i,  8, safe_float(row.get("sim_entry_px")),  bg=row_bg,
                  align="right", fmt='$#,##0.0000')
        data_cell(ws, i,  9, safe_float(row.get("sim_exit_px")),   bg=row_bg,
                  align="right", fmt='$#,##0.0000')

        pnl_col = C_POS if (sim_pnl or 0) >= 0 else C_NEG
        data_cell(ws, i, 10, sim_pnl, bg=row_bg, color=pnl_col,
                  align="right", fmt='$#,##0.00;[Red]-$#,##0.00')

        data_cell(ws, i, 11, safe_float(row.get("sim_shares")), bg=row_bg,
                  align="right", fmt="0.0")

        # Live action badge
        action_color = C_POS if live_action == "TRADED" else C_NEUTRAL
        data_cell(ws, i, 12, live_action, bg=row_bg, color=action_color,
                  bold=(live_action == "TRADED"), align="center")

        data_cell(ws, i, 13, safe_float(row.get("actual_entry_px")), bg=row_bg,
                  align="right", fmt='$#,##0.0000')
        data_cell(ws, i, 14, safe_float(row.get("actual_exit_px")),  bg=row_bg,
                  align="right", fmt='$#,##0.0000')

        apnl_col = C_POS if (act_pnl or 0) >= 0 else C_NEG
        data_cell(ws, i, 15, act_pnl, bg=row_bg, color=apnl_col,
                  align="right", fmt='$#,##0.00;[Red]-$#,##0.00')

        slip_col = C_NEG if (slip_ps or 0) > 0 else C_POS
        data_cell(ws, i, 16, slip_ps, bg=row_bg, color=slip_col if slip_ps else C_NEUTRAL,
                  align="right", fmt='+$#,##0.0000;-$#,##0.0000')
        data_cell(ws, i, 17, slip_total, bg=row_bg,
                  color=(C_NEG if (slip_total or 0) > 0 else C_POS) if slip_total else C_NEUTRAL,
                  align="right", fmt='+$#,##0.00;-$#,##0.00')

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


# ── Sheet 2: Daily Summary ────────────────────────────────────────────────────

def build_summary(wb, rows: list[dict]) -> None:
    ws = wb.create_sheet("Summary")
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 32

    headers = [
        ("Date",           11),
        ("Total Halts",    11),
        ("Qualifying",     11),
        ("Matched",        10),
        ("Missed",         10),
        ("Unexpected",     11),
        ("Match Rate",     11),
        ("Sim P&L",        10),
        ("Actual P&L",     10),
        ("Total Slippage", 13),
        ("Account Equity", 14),
    ]
    for col, (label, width) in enumerate(headers, 1):
        header_cell(ws, 1, col, label, width)

    # Group rows by date
    from collections import defaultdict
    by_date = defaultdict(list)
    for row in rows:
        by_date[row.get("date", "")].append(row)

    for i, date in enumerate(sorted(by_date.keys()), 2):
        day_rows   = by_date[date]
        bg         = C_ALT_ROW if i % 2 == 0 else C_WHITE

        total      = len(day_rows)
        qualifying = sum(1 for r in day_rows if r.get("filter_result") == "PASS")
        matched    = sum(1 for r in day_rows if r.get("filter_result") == "PASS"
                         and r.get("live_bot_action") == "TRADED")
        missed     = sum(1 for r in day_rows if r.get("filter_result") == "PASS"
                         and r.get("live_bot_action") == "SKIPPED")
        unexpected = sum(1 for r in day_rows if r.get("filter_result") == "SKIP"
                         and r.get("live_bot_action") == "TRADED")
        match_rate = matched / qualifying if qualifying > 0 else None
        sim_pnl    = sum(safe_float(r.get("sim_pnl"), 0) for r in day_rows)
        act_pnl    = sum(safe_float(r.get("actual_pnl"), 0) for r in day_rows)
        slip_total = sum(safe_float(r.get("entry_slip_total"), 0) for r in day_rows)

        # Row colour based on match quality
        if missed > 0 or unexpected > 0:
            row_bg = C_MISS_BG
        elif qualifying > 0 and matched == qualifying:
            row_bg = C_PASS_BG
        else:
            row_bg = bg

        data_cell(ws, i,  1, date,       bg=row_bg, bold=True, align="center")
        data_cell(ws, i,  2, total,      bg=row_bg, align="center")
        data_cell(ws, i,  3, qualifying, bg=row_bg, align="center")
        data_cell(ws, i,  4, matched,    bg=row_bg, align="center",
                  color=C_POS if matched > 0 else C_NEUTRAL, bold=matched > 0)
        data_cell(ws, i,  5, missed,     bg=row_bg, align="center",
                  color=C_NEG if missed > 0 else C_NEUTRAL, bold=missed > 0)
        data_cell(ws, i,  6, unexpected, bg=row_bg, align="center",
                  color="EA580C" if unexpected > 0 else C_NEUTRAL, bold=unexpected > 0)
        data_cell(ws, i,  7, match_rate, bg=row_bg, align="center", fmt="0%")
        data_cell(ws, i,  8, sim_pnl if sim_pnl else None, bg=row_bg, align="right",
                  fmt='$#,##0.00;[Red]-$#,##0.00',
                  color=C_POS if sim_pnl >= 0 else C_NEG)
        data_cell(ws, i,  9, act_pnl if act_pnl else None, bg=row_bg, align="right",
                  fmt='$#,##0.00;[Red]-$#,##0.00',
                  color=C_POS if act_pnl >= 0 else C_NEG)
        slip_col = C_NEG if slip_total > 0 else C_POS
        data_cell(ws, i, 10, slip_total if slip_total else None, bg=row_bg, align="right",
                  fmt='+$#,##0.00;-$#,##0.00', color=slip_col if slip_total else C_NEUTRAL)
        data_cell(ws, i, 11, None, bg=row_bg, align="right", fmt='$#,##0.00')

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # ── Totals row ────────────────────────────────────────────────────────────
    last = len(by_date) + 2
    ws.row_dimensions[last].height = 20
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=last, column=col)
        c.fill   = PatternFill("solid", fgColor="E2E8F0")
        c.font   = Font(bold=True, size=10, color=C_NEUTRAL)
        c.border = BORDER
        c.alignment = Alignment(horizontal="center", vertical="center")

    ws.cell(row=last, column=1).value = "TOTAL / AVG"

    n = len(by_date)
    if n > 0:
        data_row = last
        for col, letter in [(2, "B"), (3, "C"), (4, "D"), (5, "E"), (6, "F")]:
            c = ws.cell(row=data_row, column=col)
            c.value  = f"=SUM({letter}2:{letter}{last-1})"
            c.font   = Font(bold=True, size=10, color=C_NEUTRAL)
            c.fill   = PatternFill("solid", fgColor="E2E8F0")
            c.border = BORDER
            c.alignment = Alignment(horizontal="center")

        for col, letter, fmt in [
            (8,  "H", '$#,##0.00;[Red]-$#,##0.00'),
            (9,  "I", '$#,##0.00;[Red]-$#,##0.00'),
            (10, "J", '+$#,##0.00;-$#,##0.00'),
        ]:
            c = ws.cell(row=data_row, column=col)
            c.value          = f"=SUM({letter}2:{letter}{last-1})"
            c.font           = Font(bold=True, size=10, color=C_NEUTRAL)
            c.fill           = PatternFill("solid", fgColor="E2E8F0")
            c.border         = BORDER
            c.number_format  = fmt
            c.alignment      = Alignment(horizontal="right")


# ── Sheet 3: Equity Curve ─────────────────────────────────────────────────────

def build_equity(wb, rows: list[dict]) -> None:
    ws = wb.create_sheet("Equity")
    ws.freeze_panes = "A2"

    headers = [
        ("Date",           11),
        ("Actual P&L",     12),
        ("Sim P&L",        12),
        ("Cumulative Actual", 16),
        ("Cumulative Sim",    14),
        ("Slippage Cost",     13),
        ("Cum. Slippage",     13),
    ]
    for col, (label, width) in enumerate(headers, 1):
        header_cell(ws, 1, col, label, width)

    from collections import defaultdict
    by_date = defaultdict(list)
    for row in rows:
        by_date[row.get("date", "")].append(row)

    cum_actual  = 0.0
    cum_sim     = 0.0
    cum_slip    = 0.0

    for i, date in enumerate(sorted(by_date.keys()), 2):
        day_rows  = by_date[date]
        bg        = C_ALT_ROW if i % 2 == 0 else C_WHITE
        act_pnl   = sum(safe_float(r.get("actual_pnl"), 0) for r in day_rows)
        sim_pnl   = sum(safe_float(r.get("sim_pnl"), 0) for r in day_rows)
        slip      = sum(safe_float(r.get("entry_slip_total"), 0) for r in day_rows)

        cum_actual += act_pnl
        cum_sim    += sim_pnl
        cum_slip   += slip

        data_cell(ws, i, 1, date,       bg=bg, bold=True, align="center")
        data_cell(ws, i, 2, act_pnl  or None, bg=bg, align="right",
                  fmt='$#,##0.00;[Red]-$#,##0.00',
                  color=C_POS if act_pnl >= 0 else C_NEG)
        data_cell(ws, i, 3, sim_pnl  or None, bg=bg, align="right",
                  fmt='$#,##0.00;[Red]-$#,##0.00',
                  color=C_POS if sim_pnl >= 0 else C_NEG)
        data_cell(ws, i, 4, cum_actual or None, bg=bg, align="right",
                  fmt='$#,##0.00;[Red]-$#,##0.00', bold=True,
                  color=C_POS if cum_actual >= 0 else C_NEG)
        data_cell(ws, i, 5, cum_sim    or None, bg=bg, align="right",
                  fmt='$#,##0.00;[Red]-$#,##0.00', bold=True,
                  color=C_POS if cum_sim >= 0 else C_NEG)
        data_cell(ws, i, 6, slip      or None, bg=bg, align="right",
                  fmt='+$#,##0.00;-$#,##0.00',
                  color=C_NEG if slip > 0 else C_POS if slip < 0 else C_NEUTRAL)
        data_cell(ws, i, 7, cum_slip  or None, bg=bg, align="right",
                  fmt='+$#,##0.00;-$#,##0.00', bold=True,
                  color=C_NEG if cum_slip > 0 else C_POS if cum_slip < 0 else C_NEUTRAL)

    n_rows = len(by_date)
    if n_rows >= 2:
        # Cumulative P&L chart
        chart = LineChart()
        chart.title   = "Cumulative P&L — Actual vs Sim (no slippage)"
        chart.style   = 10
        chart.y_axis.title = "P&L ($)"
        chart.x_axis.title = "Trading Day"
        chart.width   = 26
        chart.height  = 14

        actual_data = Reference(ws, min_col=4, min_row=1, max_row=n_rows + 1)
        sim_data    = Reference(ws, min_col=5, min_row=1, max_row=n_rows + 1)

        chart.add_data(actual_data, titles_from_data=True)
        chart.add_data(sim_data,    titles_from_data=True)

        chart.series[0].graphicalProperties.line.solidFill  = "2563EB"
        chart.series[0].graphicalProperties.line.width      = 20000
        chart.series[1].graphicalProperties.line.solidFill  = "10B981"
        chart.series[1].graphicalProperties.line.width      = 20000
        chart.series[1].graphicalProperties.line.dashDot    = "dash"

        ws.add_chart(chart, "I2")


# ── Sheet 4: Slippage Analysis ────────────────────────────────────────────────

def build_slippage(wb, rows: list[dict]) -> None:
    """Only rows where live bot actually traded."""
    traded = [r for r in rows if r.get("live_bot_action") == "TRADED"]
    if not traded:
        return

    ws = wb.create_sheet("Slippage")
    ws.freeze_panes = "A2"

    headers = [
        ("Date",          11),
        ("Symbol",         9),
        ("Resume",         9),
        ("Sim Open",      10),
        ("Actual Fill",   10),
        ("Slip/Share",    10),
        ("Shares",         9),
        ("Slip Total",    11),
        ("Sim P&L",       10),
        ("Actual P&L",    10),
        ("P&L Diff",      10),
    ]
    for col, (label, width) in enumerate(headers, 1):
        header_cell(ws, 1, col, label, width)

    for i, row in enumerate(traded, 2):
        bg        = C_ALT_ROW if i % 2 == 0 else C_WHITE
        sim_e     = safe_float(row.get("sim_entry_px"))
        act_e     = safe_float(row.get("actual_entry_px"))
        shares    = safe_float(row.get("sim_shares"))
        slip_ps   = safe_float(row.get("entry_slip_per_sh"))
        slip_tot  = safe_float(row.get("entry_slip_total"))
        sim_pnl   = safe_float(row.get("sim_pnl"))
        act_pnl   = safe_float(row.get("actual_pnl"))
        pnl_diff  = (act_pnl - sim_pnl) if (act_pnl is not None and sim_pnl is not None) else None

        slip_bg   = "FEE2E2" if (slip_tot or 0) > 0 else "D1FAE5" if (slip_tot or 0) < 0 else bg

        data_cell(ws, i,  1, row.get("date"),    bg=bg, align="center")
        data_cell(ws, i,  2, row.get("symbol"),  bg=bg, bold=True)
        data_cell(ws, i,  3, row.get("resume_time"), bg=bg, align="center")
        data_cell(ws, i,  4, sim_e,   bg=bg,      align="right", fmt='$#,##0.0000')
        data_cell(ws, i,  5, act_e,   bg=bg,      align="right", fmt='$#,##0.0000')
        data_cell(ws, i,  6, slip_ps, bg=slip_bg, align="right",
                  fmt='+$#,##0.0000;-$#,##0.0000',
                  color=C_NEG if (slip_ps or 0) > 0 else C_POS if (slip_ps or 0) < 0 else C_NEUTRAL)
        data_cell(ws, i,  7, shares,   bg=bg, align="right", fmt="0.0")
        data_cell(ws, i,  8, slip_tot, bg=slip_bg, align="right",
                  fmt='+$#,##0.00;-$#,##0.00', bold=True,
                  color=C_NEG if (slip_tot or 0) > 0 else C_POS if (slip_tot or 0) < 0 else C_NEUTRAL)
        data_cell(ws, i,  9, sim_pnl,  bg=bg, align="right",
                  fmt='$#,##0.00;[Red]-$#,##0.00',
                  color=C_POS if (sim_pnl or 0) >= 0 else C_NEG)
        data_cell(ws, i, 10, act_pnl,  bg=bg, align="right",
                  fmt='$#,##0.00;[Red]-$#,##0.00',
                  color=C_POS if (act_pnl or 0) >= 0 else C_NEG)
        data_cell(ws, i, 11, pnl_diff, bg=bg, align="right",
                  fmt='+$#,##0.00;-$#,##0.00', bold=True,
                  color=C_NEG if (pnl_diff or 0) < 0 else C_POS if (pnl_diff or 0) > 0 else C_NEUTRAL)

    # Totals
    last = len(traded) + 2
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=last, column=col)
        c.fill   = PatternFill("solid", fgColor="E2E8F0")
        c.font   = Font(bold=True, size=10, color=C_NEUTRAL)
        c.border = BORDER
        c.alignment = Alignment(horizontal="center")
    ws.cell(row=last, column=1).value = "TOTAL"
    for col, letter, fmt in [
        (8,  "H", '+$#,##0.00;-$#,##0.00'),
        (9,  "I", '$#,##0.00;[Red]-$#,##0.00'),
        (10, "J", '$#,##0.00;[Red]-$#,##0.00'),
        (11, "K", '+$#,##0.00;-$#,##0.00'),
    ]:
        c = ws.cell(row=last, column=col)
        c.value         = f"=SUM({letter}2:{letter}{last-1})"
        c.font          = Font(bold=True, size=10, color=C_NEUTRAL)
        c.fill          = PatternFill("solid", fgColor="E2E8F0")
        c.border        = BORDER
        c.number_format = fmt
        c.alignment     = Alignment(horizontal="right")

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print(f"Loading {CSV_PATH} ...")
    rows = load_csv()
    print(f"  {len(rows):,} rows loaded")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # remove default empty sheet

    print("Building Daily Log sheet ...")
    build_daily_log(wb, rows)

    print("Building Summary sheet ...")
    build_summary(wb, rows)

    print("Building Equity sheet ...")
    build_equity(wb, rows)

    print("Building Slippage sheet ...")
    build_slippage(wb, rows)

    wb.save(OUTPUT_PATH)
    print(f"\n✓ Report saved to {OUTPUT_PATH}")
    print(f"\nTo download on Windows:")
    print(f'  scp trade-server:/root/Live_Trader_Halts/bot/{OUTPUT_PATH} .')


if __name__ == "__main__":
    main()
